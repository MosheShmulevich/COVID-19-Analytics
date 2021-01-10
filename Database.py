from openpyxl import *

Covid19DB = load_workbook(filename='Database.xlsx', data_only=True)  # Covid19 database
MainSheet = Covid19DB['Cities Data']
CityDB = ((load_workbook(filename="Israel city list.xlsx"))['Israel Cities'])['A']  # Israel cities sheet column A
IL_CityList = [CityDB[x].value for x in range(len(CityDB))]

RecoveredDB = load_workbook(filename='RecoveredDatabase.xlsx', data_only=True)  # Recovered patients database

ResultDB = load_workbook(filename='TestDatabase.xlsx', data_only=True)  # Covid19 Tests results database

userDb = load_workbook(filename='UserDatabase.xlsx')  # System users database
userSheet = userDb['users']  # UserDatabase users sheet



