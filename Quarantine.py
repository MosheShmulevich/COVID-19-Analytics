import openpyxl

wb = openpyxl.load_workbook('quarantine.xlsx')
print("Enter the patients city:")
city = input()
sheet = wb.get_sheet_by_name(city)
print("Enter patient's name:")
name = input()
print("Enter patient's birth date:")
birth_date = input()
print("Enter patient's ID:")
id1 = input()
print("Enter patient's last test date:")
test_date = input()
print("Enter patient's last test result:")
test_result = input()
print("Enter patient's quarantine place: (home, corona motel, hospital)")
place = input()
row1 = 0
for cell in sheet["A"]:    # find the empty raw
    if cell.value is None:
        row1 = cell.row
        break
    else:
        row1 = cell.row + 1
sheet.cell(row=row1, column=1).value = name    # inputs to excel
sheet.cell(row=row1, column=2).value = birth_date
sheet.cell(row=row1, column=3).value = id1
sheet.cell(row=row1, column=4).value = test_date
sheet.cell(row=row1, column=5).value = test_result
sheet.cell(row=row1, column=6).value = place


wb.save('quarantine_update.xlsx')   # save