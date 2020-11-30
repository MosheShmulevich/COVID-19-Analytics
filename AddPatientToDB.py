import Database
from dataclasses import dataclass
from datetime import *
from openpyxl import load_workbook


@dataclass()
class New_Patient:
    firstname: str
    lastname: str
    ID: int
    birth_date: datetime
    test_date: datetime
    test_result: str
    is_tested: str
    status: str


def CreateNewSheet(city):  # function for creating a new 'city' sheet in xlsx database
    def FormatCells(city_sheet):
        for row in range(2, 28):
            for column in range(3, 7):
                if column == 3:
                    city_sheet.cell(row, column)

    def SwitchTitle(column):  # function for adding to the xlsx table columns, titles
        switcher = {
            1: 'Firstname',
            2: 'Lastname',
            3: 'ID',
            4: 'Birth date',
            5: 'Test date',
            6: 'Patient Status'}
        return switcher

    Database.Covid19DB.create_sheet(city)  # creating a new sheet with 'city' name
    city_sheet = Database.Covid19DB[city]  # implementing the 'city' sheet into a variable "city_sheet"
    for colmn in range(1, 7):  # adding to the xlsx table columns, titles
        city_sheet.cell(row=1, column=colmn).value = SwitchTitle(colmn)[colmn]


def addPatient(New_Patient, city):  # function for adding a patient to the 'city' sheet
    if city not in Database.Covid19DB.sheetnames:  # if there is no 'city' sheet , calls a fucntion to create one
        CreateNewSheet(city)

    New_Patient.is_tested = None
    city_sheet = Database.Covid19DB[city]

    def switch_input(k):  # function to choose patient's parameters
        switcher = {
            1: New_Patient.firstname,
            2: New_Patient.lastname,
            3: New_Patient.ID,
            4: New_Patient.birth_date,
            5: New_Patient.test_date,
            6: New_Patient.status
        }
        return switcher

    for row in range(2, 28):  # starts from row#2 because #1 is titles and end in #27(temporarly)
        for j in range(1, 2):  # start in column#1 to check if its empty
            if city_sheet.cell(row=row, column=j).value is None:
                for col in range(1, 7):  # adding patient parameters into the empty row by columns
                    city_sheet.cell(row=row, column=col).value = switch_input(col)[col]
                Database.Covid19DB.save('Database.xlsx')  # after adding parameters,the system "autosaves" the xlsx file
                break
            else:
                break
        if New_Patient.firstname == city_sheet.cell(row=row,
                                                    column=1).value:  # after adding a patient, stops the outer loop
            break
