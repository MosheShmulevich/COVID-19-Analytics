from openpyxl import load_workbook
import xlrd
from xlwings import sheets

workbook = xlrd.open_workbook('labTests.xlsx')  # worker in the lab writing this excel
worksheet = workbook.sheet_by_name('Sheet1')
z = 1
city = worksheet.cell(z, 3)  # column 3 row z
i = 1

while i is not None:  # until the end of the row
    test_res = worksheet.cell(i, 6)
    numOfTest = worksheet.cell(i, 5)  # how many times this patient had a test
    if test_res == "positive":  # result  positive to corona
        wb = load_workbook('Database.xlsx.')
        Sheet1 = wb[sheets[city]]  # the sheets are by cities
        Sheet1.cell(row=i, column=5).value = "Covid confirmed"
    elif test_res == "negative":
        if numOfTest > 0:
            wb = load_workbook('Database.xlsx')
            Sheet1 = wb[sheets[0]]
            Sheet1.cell(row=i, column=5).value = "Recovered"
    i = i + 1
