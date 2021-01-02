import webbrowser
import openpyxl


def choose():
    print("To view Covid 19 cases map enter 1, To view Covid 19 cases for a specific city enter 2:")
    choice1 = int(input())
    if choice1 == 1:
        map1()
    if choice1 == 2:
        by_city()
    else:
        print("to try again enter 3 , to exit enter 0 :")
        choice2 = int(input())
        if choice2 == 3:
            choose()
        if choice2 == 0:
            exit()


def map1():
    webbrowser.open('https://www.govmap.gov.il/sites/coronamap.html')


def by_city():
    print('Enter a city:')
    city = input()
    wb = openpyxl.load_workbook(filename='confirmed_patients.xlsx')
    sheet = wb[city]
    last_row = sheet.max_row
    print(last_row)
    confirmed_count = last_row - 1
    print('in ', city)
    print("total cases: ", confirmed_count)

    wb = openpyxl.load_workbook(filename='recovered.xlsx')
    sheet = wb[city]
    last_row = sheet.max_row
    print(last_row)
    recovered_count = last_row - 2
    print("recovered: ", recovered_count)

    wb = openpyxl.load_workbook(filename='death_cases.xlsx')
    sheet = wb[city]
    last_row = sheet.max_row
    death_count = last_row - 2
    print("death cases: ", death_count)


choose()






