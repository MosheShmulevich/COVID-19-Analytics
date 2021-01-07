from tkinter import *
import openpyxl


def to_excel():
    file_name = 'volunteer.xlsx'
    wb = openpyxl.load_workbook(file_name)
    city = clicked.get()
    sheet = wb[city]
    name = textentry1.get()
    phone = textentry2.get()
    date = clicked1.get()
    row1 = 0
    for cell in sheet["A"]:  # find the empty row
        if cell.value is None:
            row1 = cell.row
            break
        else:
            row1 = cell.row + 1
    sheet.cell(row=row1, column=1).value = name
    sheet.cell(row=row1, column=2).value = phone
    sheet.cell(row=row1, column=3).value = date
    wb.save(file_name)
    exit()


window = Tk()
window.title("Volunteer sign up")
photo_map = PhotoImage(file="map.png")
Label(window, image=photo_map, bg="black").grid(row=0, column=1, sticky=W)
Label(window, text="Dear volunteer ",bg="white",fg="black",font=" none 12 bold").grid(row=0, column=0,sticky=W)
Label(window, text="every city with a black star",bg="white",fg="black",font=" none 12 bold").grid(row=1, column=0,sticky=W)
Label(window, text=" has a volunteer place",bg="white",fg="black",font=" none 12 bold").grid(row=2, column=0,sticky=W)
Label(window, text="Sign in:",bg="white",fg="black",font=" none 12 bold").grid(row=4, column=0,sticky=W)
Label(window, text="Name:",bg="white",fg="black",font=" none 12").grid(row=5, column=0,sticky=W)
textentry1 = Entry(window, width=15, bg="white")
textentry1.grid(row=5, column=1)
Label(window, text="Phone number:",bg="white",fg="black",font=" none 12").grid(row=6, column=0,sticky=W)
textentry2 = Entry(window, width=15, bg="white")
textentry2.grid(row=6, column=1)
Label(window, text="Date to volunteer:",bg="white",fg="black",font=" none 12").grid(row=7, column=0,sticky=W)
dates = ["10.1.20", "11.1.20", "12.1.20", "13.1.20","14.1.20","17.1.20","18.1.20","19.1.20"]
clicked1 = StringVar()
clicked1.set(dates[0])
drop1 = OptionMenu(window,clicked1,*dates)
drop1.grid(row=7,column=1)
Label(window, text="City to volunteer:",bg="white",fg="black",font=" none 12").grid(row=8, column=0,sticky=W)
cities = ["Beer Sheva", "Tel Aviv", "Mizpe Ramon", "Heifa","Afula","Hadera","Jerusalem","Eilat"]
clicked = StringVar()
clicked.set(cities[0])
drop = OptionMenu(window,clicked,*cities)
drop.grid(row=8,column=1)
Button(window, text="SUBMIT",width=6,command = to_excel).grid(row=9, column=1)


window.mainloop()








