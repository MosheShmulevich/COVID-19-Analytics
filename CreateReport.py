from openpyxl import *
import Database


def create_city_report(city):
    for row in range(2, 100):
        if Database.city_sheet.cell(row=row, column=1).value == city:
            city_row = row
    print("Here are the data for the city:", city)
    for column in range(2, 14):
        print(Database.city_sheet.cell(row=1, column=column).value, "-",
              Database.city_sheet.cell(row=city_row, column=column).value)


def create_CityAndOption_report():
    def OptionChoose(option):
        switcher = {
            1: 'Total Cases',
            2: 'New Cases',
            3: 'Active Cases',
            4: 'Total Deaths',
            5: 'New Deaths',
            6: 'Total Recovered',
            7: 'New Recovered',
            8: 'Total Tests',
            9: 'Cases to 1M Population',
            10: 'Tests to 1M Population',
            11: 'Deaths to 1M Population',
            12: 'Population'
        }
        return switcher

    option = input("Which data would you like to see? (Choose number)\n1 - Total Cases\n2 - New Cases\n3 - Active "
                   "Cases\n4 - "
                   "Total Deaths\n5 - New Deaths\n6 - Total Recovered\n "
                   "7 - New Recovered\n8 - Total Tests\n9 - Cases to 1M Population\n10 - Tests to 1M Population\n"
                   "11 - Deaths to 1M Population\n12 - Population\n")
    for row in range(2, 100):
        for column in range(2, 14):
            if Database.city_sheet.cell(row=1, column=column) == OptionChoose(option):
                print(Database.city_sheet.cell(row=row, column=1).value, ":",
                      Database.city_sheet.cell(row=row, column=column).value)

create_CityAndOption_report()