from tkinter import *
import openpyxl


def to_xl(list1):
    profession = list1[0]
    name = list1[1]
    phone = list1[2]
    city = list1[3]
    file_name = "profession.xlsx"
    wb = openpyxl.load_workbook(file_name)
    sheet = wb.active
    row1 = 0
    for cell in sheet["A"]:  # find the empty row
        if cell.value is None:
            row1 = cell.row
            break
        else:
            row1 = cell.row + 1
    sheet.cell(row=row1, column=1).value = profession
    sheet.cell(row=row1, column=2).value = name
    sheet.cell(row=row1, column=3).value = phone
    sheet.cell(row=row1, column=4).value = city
    wb.save(file_name)
    exit()


def help1():
    window1 = Tk()
    window1.configure(width=500, height=500, background="gray4")
    Label(window1, text="After pressing 'send' our worker will contact you: ", bg="gray4", fg="white", font="none 12 bold").grid(row=0, column=0)
    Label(window1, text="Enter your profession: ", bg="gray4", fg="white", font=" none 12 bold").grid(row=1, column=0)
    textentry = Entry(window1, width=15, bg="white")
    textentry.grid(row=1, column=1)
    Label(window1, text="Enter your name: ", bg="gray4", fg="white", font=" none 12 bold").grid(row=2, column=0)
    textentry1 = Entry(window1, width=15, bg="white")
    textentry1.grid(row=2, column=1)
    Label(window1, text="Enter your phone number: ", bg="gray4", fg="white", font=" none 12 bold").grid(row=3, column=0)
    textentry2 = Entry(window1, width=15, bg="white")
    textentry2.grid(row=3, column=1)
    Label(window1, text="Enter the city: ", bg="gray4", fg="white", font=" none 12 bold").grid(row=4, column=0)
    textentry3 = Entry(window1, width=15, bg="white")
    textentry3.grid(row=4, column=1)
    Button(window1, text="SAVE", width=4, command=lambda: save()).grid(row=6, column=0)
    def save():
        profession = textentry.get()
        name = textentry1.get()
        phone = textentry2.get()
        city = textentry3.get()
        list1 = [profession, name, phone, city]
        Button(window1, text="SEND", width=4, command=lambda:to_xl(list1)).grid(row=6, column=0)


def show_info():
    option = clicked.get()
    if option == "Covid19 ":
        window1 = Tk()
        window1.configure(width=500,height=500,background="gray4")
        Label(window1,text="COVID-19 is the disease caused by a new coronavirus called SARS-CoV-2.  WHO first learned ",bg="gray4",fg="white",
              font="none 12").grid(row=0)
        Label(window1,text="of ‘viral pneumonia’ in Wuhan, People’s Republic of China.",bg="gray4",fg="white",font="font 12").grid(row=1)
        Label(window1, text="this new virus on 31 December 2019, following a report of a cluster of cases ", bg="gray4", fg="white",
              font="font 12").grid(row=2)
        Label(window1, text="The most common symptoms of COVID-19 are: Fever, Dry Cough,Fatigue.People of all ages who ", bg="gray4", fg="white",
              font="font 12").grid(row=3)
        Label(window1, text="experience fever and/or cough associated with difficulty breathing or shortness of breath,", bg="gray4", fg="white",
              font="font 12").grid(row=4)
        Label(window1, text="chest pain or pressure,", bg="gray4", fg="white",font="font 12").grid(row=5)
        Label(window1, text="or loss of speech or movement should seek medical care immediately. If possible, call your"
              , bg="gray4", fg="white", font="font 12").grid(row=6)
        Label(window1, text="health care provider, hotline or health facility first, so you can be directed to the right clinic", bg="gray4", fg="white", font="font 12").grid(row=7)
        Label(window1, text="Among those who develop symptoms, most (about 80%) recover from the disease without needing", bg="gray4", fg="white", font="font 12").grid(row=8)
        Label(window1, text="hospital treatment. About 15% become seriously ill and require oxygen and 5% become", bg="gray4", fg="white", font="font 12").grid(row=9)
        Label(window1, text="critically ill and need intensive care", bg="gray4", fg="white", font="font 12").grid(row=10)
        Label(window1, text="Complications leading to death may include respiratory failure, acute respiratory distress", bg="gray4", fg="white", font="font 12").grid(row=11)
        Label(window1, text="syndrome (ARDS), sepsis and septic shock, thromboembolism, and/or multiorgan failure", bg="gray4", fg="white", font="font 12").grid(
            row=12)
        Label(window1, text="including injury of the heart, liver or kidneys.", bg="gray4", fg="white", font="font 12").grid(row=13)
        Label(window1, text="People aged 60 years and over, and those with underlying medical problems like high blood", bg="gray4", fg="white",
              font="font 12").grid(row=14)
        Label(window1, text="pressure, heart and lung problems, diabetes, obesity or cancer, are at higher risk of", bg="gray4", fg="white",
              font="font 12").grid(row=15)
        Label(window1, text="developing serious illness.", bg="gray4", fg="white",
              font="font 12").grid(row=16)
        Label(window1, text="WHO is working with our Global Technical Network for Clinical Management of COVID-19,", bg="gray4", fg="white",
              font="font 12").grid(row=17)
        Label(window1, text="researchers and patient groups around the world to design and carry out studies of ", bg="gray4", fg="white",
              font="font 12").grid(row=18)
        Label(window1, text="researchers and patient groups around the world to design and carry out studies of patients", bg="gray4", fg="white",
              font="font 12").grid(row=19)
        Label(window1, text="beyond the initial acute course of illness to understand the proportion of patients who", bg="gray4", fg="white", font="font 12").grid(row=20)
        Label(window1, text="have long term effects, how long they persist, and why they occur.  These studies will", bg="gray4", fg="white",
              font="font 12").grid(row=21)
        Label(window1, text="be used to develop further guidance for patient care.",bg="gray4",fg="white",
              font="font 12").grid(row=22)

    if option == " guidelines":
        window1 = Tk()
        window1.configure(width=500, height=500, background="gray4")
        Label(window1,text="If COVID-19 is spreading in your community, stay safe by taking some simple precautions,",bg="gray4",fg="white",font="font 12").grid(row=0)
        Label(window1, text="such as physical distancing, wearing a mask, keeping rooms well ventilated,",bg="gray4", fg="white", font="font 12").grid(row=1)
        Label(window1, text="avoiding crowds, cleaning your hands, and coughing into a bent elbow or tissue.", bg="gray4",fg="white", font="font 12").grid(row=2)
        Label(window1, text="Check local advice where you live and work. Do it all!",bg="gray4", fg="white", font="font 12").grid(row=3)
        Label(window1, text="-Maintain at least a 1-metre distance between yourself and others to reduce your risk of infection when they cough, sneeze or speak", bg="gray4", fg="white",
              font="font 12").grid(row=4)
        Label(window1, text="Maintain an even greater distance between yourself and others when indoors.", bg="gray4", fg="white",
              font="font 12").grid(row=5)
        Label(window1, text="The further away, the better.", bg="gray4",fg="white",font="font 12").grid(row=6)
        Label(window1, text="-Make wearing a mask a normal part of being around other people.", bg="gray4", fg="white", font="font 12").grid(row=7)
        Label(window1, text="The appropriate use, storage and cleaning or disposal are essential to make masks as effective as possible.", bg="gray4", fg="white", font="font 12").grid(row=8)
        Label(window1, text="Here are the basics of how to wear a mask:", bg="gray4", fg="white", font="font 12").grid(row=9)
        Label(window1, text="-Clean your hands before you put your mask on, as well as before and after you take it off, and after you touch it at any time.", bg="gray4", fg="white", font="font 12").grid(row=10)
        Label(window1, text="-Make sure it covers both your nose, mouth and chin.", bg="gray4", fg="white", font="font 12").grid(row=11)
        Label(window1, text="-When you take off a mask, store it in a clean plastic bag, and every day either wash it if it’s a fabric mask, or dispose of a medical mask in a trash bin.", bg="gray4", fg="white", font="font 12").grid(row=12)
        Label(window1, text="-Make sure it covers both your nose, mouth and chin.", bg="gray4", fg="white",
              font="font 12").grid(row=13)
        Label(window1, text="-When you take off a mask, store it in a clean plastic bag, and every day either wash it if it’s a fabric mask,", bg="gray4", fg="white", font="font 12").grid(row=14)
        Label(window1, text=",or dispose of a medical mask in a trash bin.", bg="gray4", fg="white", font="font 12").grid(row=15)
        Label(window1, text="-Don’t use masks with valves.", bg="gray4", fg="white",font="font 12").grid(row=16)

    if option == "vaccination":
        window1 = Tk()
        window1.configure(width=500, height=500, background="gray4")
        Label(window1, text="There are three COVID-19 vaccines for which certain national regulatory authorities have authorized the use.", bg="gray4", fg="white", font="font 12").grid(row=0)
        Label(window1, text="None have yet received WHO EUL/PQ authorization but we expect an assessment on the Pfizer vaccine by the end of December and for some other candidates soon thereafter. ",bg="gray4", fg="white", font="font 12").grid(row=1)
        Label(window1, text="Once vaccines are demonstrated to be safe and efficacious, they must be approved by national regulators, manufactured to exacting standards, and distributed.", bg="gray4", fg="white", font="font 12").grid(row=2)
        Label(window1,text="WHO is working with partners around the world to help coordinate key steps in this process, ",bg="gray4", fg="white", font="font 12").grid(row=3)
        Label(window1,text=", including to facilitate equitable access to safe and effective COVID-19 vaccines for the billions of people who will need them.",bg="gray4", fg="white", font="font 12").grid(row=4)
        Label(window1,text="The impact of COVID-19 vaccines on the pandemic will depend on several factors.",bg="gray4", fg="white", font="font 12").grid(row=5)
        Label(window1,text="These include factors such as the effectiveness of the vaccines; how quickly they are approved, manufactured, and delivered; and how many people get vaccinated.",
              bg="gray4", fg="white", font="font 12").grid(row=6)
        Label(window1, text="Most scientists anticipate that, like most other vaccines, COVID-19 vaccines will not be 100% effective.",bg="gray4", fg="white", font="font 12").grid(row=7)
        Label(window1, text="WHO is working to help ensure that any approved vaccines are as effective as possible, so they can have the greatest impact on the pandemic.",bg="gray4", fg="white", font="font 12").grid(row=8)
        Label(window1, text="Most scientists anticipate that, like most other vaccines, COVID-19 vaccines will not be 100% effective.",bg="gray4", fg="white", font="font 12").grid(row=9)
        Label(window1,text="Pfizer Inc.",bg="gray4", fg="white", font="font 12 bold").grid(row=10)
        Label(window1, text="235 East 42nd Street NY, NY 10017", bg="gray4", fg="white", font="font 12 bold").grid(row=11)
        Label(window1, text="(212) 733-2323", bg="gray4", fg="white", font="font 12 bold").grid(row=12)

    if option == "mental health":
        window1 = Tk()
        window1.configure(width=500, height=500, background="gray4")
        Label(window1, text="Fear, worry, and stress are normal responses to perceived or real threats, and at times when we are faced with uncertainty or the unknown.", bg="gray4", fg="white", font="font 12").grid(row=0)
        Label(window1, text="So it is normal and understandable that people are experiencing fear in the context of the COVID-19 pandemic.", bg="gray4", fg="white", font="font 12").grid(row=1)
        Label(window1,
              text="Added to the fear of contracting the virus in a pandemic such as COVID-19 are the significant changes to our daily lives as our movements are restricted in support of efforts to contain and slow down the spread of the virus. ",
              bg="gray4", fg="white", font="font 12").grid(row=2)
        Label(window1, text="Faced with new realities of working from home, temporary unemployment, home-schooling of children, and lack of physical contact with other family members,", bg="gray4", fg="white", font="font 12").grid(row=3)
        Label(window1, text="friends and colleagues, it is important that we look after our mental, as well as our physical, health.", bg="gray4", fg="white", font="font 12").grid(row=4)
        Label(window1, text="If you need help please contact : ", bg="gray4", fg="white", font="font 12 ").grid(row=5)
        Label(window1, text="Doctor Israel Israeli ", bg="gray4", fg="white", font="font 12 ").grid(row=6)
        Label(window1, text="053213445", bg="gray4", fg="white", font="font 12 ").grid(row=7)

    if option == "blood donation":
        window1 = Tk()
        window1.configure(width=500, height=500, background="gray4")
        Label(window1, text="Are you fully recovered from a verified COVID-19 diagnosis? ", bg="gray4", fg="white", font="font 12 ").grid(row=0)
        Label(window1, text="If so, the plasma in your blood may contain COVID-19 antibodies that can attack the virus.", bg="gray4", fg="white",font="font 12 ").grid(row=1)
        Label(window1,text="This convalescent plasma is being evaluated as a possible treatment for currently ill COVID-19 patients, so your donation could help save the lives of patients battling this disease!",bg="gray4", fg="white", font="font 12 ").grid(row=2)
        Label(window1, text="The American Red Cross is testing all blood, platelet and plasma donations for COVID-19 antibodies.", bg="gray4", fg="white",font="font 12 ").grid(row=3)
        Label(window1, text="As part of that effort, plasma from whole blood donations that test positive for ", bg="gray4", fg="white",font="font 12 ").grid(row=4)
        Label(window1, text="COVID-19 antibodies may now help current coronavirus patients in need of convalescent plasma transfusions.", bg="gray4", fg="white",font="font 12").grid(row=5)
        Label(window1, text="We are committed to helping others in meaningful ways during this pandemic.", bg="gray4", fg="white",font="font 12 ").grid(row=6)
        Label(window1, text="For donation please contact:", bg="gray4", fg="white",font="font 12 bold").grid(row=7)
        Label(window1, text="08-6400138", bg="gray4", fg="white", font="font 12 bold").grid(row=8)
        Label(window1, text="Blood Bank", bg="gray4", fg="white", font="font 12 bold").grid(row=9)
        Label(window1, text="Soroka Hospital", bg="gray4", fg="white", font="font 12 bold").grid(row=10)

    if option == "professional help":
        window1 = Tk()
        window1.configure(width=500, height=500, background="gray4")
        Label(window1, text="Psychologist – Tel Aviv: ", bg="gray4", fg="white", font="font 12 bold").grid(row=0)
        Label(window1, text="Dr. Abal Horoviz . 086579487", bg="gray4", fg="white", font="font 12").grid(row=1)
        Label(window1, text="Dr. Nina Rosen . 086574321", bg="gray4", fg="white", font="font 12").grid(row=2)
        Label(window1, text="Jerusalem:", bg="gray4", fg="white", font="font 12 bold").grid(row=3)
        Label(window1, text="Dr. Rinat Cohen . 086579876", bg="gray4", fg="white", font="font 12").grid(row=4)
        Label(window1, text="Dr. Izhak Yaacov . 086554327", bg="gray4", fg="white", font="font 12").grid(row=5)
        Label(window1, text="Dogs Pensions :  - Beer Sheva:", bg="gray4", fg="white", font="font 12 bold").grid(row=6)
        Label(window1, text="Dogs world . 0543198764", bg="gray4", fg="white", font="font 12 bold").grid(row=7)
        Label(window1, text="Tel Aviv :", bg="gray4", fg="white", font="font 12 bold").grid(row=8)
        Label(window1, text="Pets Dream . 0556784356", bg="gray4", fg="white", font="font 12 ").grid(row=9)
        Label(window1, text="IF YOU WANT YOUR AD HERE PLEASE CLICK ",bg="gray4", fg="white", font="font 12 ").grid(row=10)
        Button(window1, text="HERE", width=4, command=help1).grid(row=10, column=1)

    if option == "system":
        window1 = Tk()
        window1.configure(width=500, height=500, background="gray4")
        Label(window1, text="About this system:", bg="gray4", fg="white",font="font 12 bold ").grid(row=0)
        Label(window1, text="This system includes all the analytics and information about the Covid-19 virus.", bg="gray4", fg="white", font="font 12").grid(row=1)
        Label(window1, text="The system has: patients database , recovered patients database, graphs and tables,", bg="gray4", fg="white", font="font 12").grid(row=2)
        Label(window1, text="information menu including :information about the guidelines, pandemic, vaccines  and more…", bg="gray4", fg="white", font="font 12").grid(row=3)
        Label(window1, text="The lab workers input the tests result in the system and the databases upgrade automatically after the managers approve.", bg="gray4", fg="white", font="font 12").grid(row=4)
        Label(window1, text="The system has tracing functions to avoid mistakes. ", bg="gray4", fg="white", font="font 12").grid(row=5)
        Label(window1, text="The users can see the analytics in graphs, tables etc.", bg="gray4", fg="white", font="font 12").grid(row=6)
        Label(window1, text="This system was build in Python language via “Pycharm Community”.", bg="gray4", fg="white", font="font 12").grid(row=7)
        Label(window1, text="All the information provided in this system is WHO (world health organization) approved.", bg="gray4", fg="white", font="font 12").grid(row=8)
        Label(window1, text="development team:", bg="gray4", fg="white", font="font 12 bold ").grid(row=9)
        Label(window1, text="Moshe Shmulevich", bg="gray4", fg="white", font="font 12 ").grid(row=10)
        Label(window1, text="Max Shapira", bg="gray4", fg="white", font="font 12  ").grid(row=11)
        Label(window1, text="Tanya Zelenko", bg="gray4", fg="white", font="font 12").grid(row=12)
        Label(window1, text="Amei Michaelov", bg="gray4", fg="white", font="font 12 ").grid(row=13)
        Label(window1, text="Project 35 company", bg="gray4", fg="white", font="font 12 bold ").grid(row=14)

    if option =="developer company":
        window1 = Tk()
        window1.configure(width=500, height=500, background="gray4")
        Label(window1, text="A new startup company undertaken by an entrepreneur to seek, develop, and validate a scalable economic model.", bg="gray4", fg="white", font="font 12").grid(row=0)
        Label(window1, text="The company has 4 co-founders: Moshe Shmulevich, Max Shapira, Tanya Zelenko, Amei Michaelov.", bg="gray4", fg="white", font="font 12 ").grid(row=1)
        Label(window1, text="Our principles:", bg="gray4", fg="white", font="font 14 bold ").grid(row=2)
        Label(window1, text="Lean startup", bg="gray4", fg="white", font="font 12 bold ").grid(row=3)
        Label(window1, text="Lean startup is a clear set of principles to create and design startups under limited resources.", bg="gray4", fg="white", font="font 12").grid(row=4)
        Label(window1, text="The empirical test is to de/validate these assumptions and to get an engaged understanding of the business model of the new ventures. ", bg="gray4", fg="white", font="font 12").grid(row=6)
        Label(window1, text="Hence, lean startup is a set of principles for entrepreneurial learning and business model design.", bg="gray4", fg="white", font="font 12").grid(row=7)
        Label(window1,text="More precisely, it is a set of design principles aimed for iteratively experiential learning under uncertainty in an engaged empirical manner. ",bg="gray4", fg="white", font="font 12").grid(row=8)
        Label(window1, text="Typically, lean startup focuses on a few lean principles:", bg="gray4", fg="white", font="font 12 bold ").grid(row=9)
        Label(window1, text="-find a problem worth solving, then define a solution", bg="gray4", fg="white",font="font 12").grid(row=10)
        Label(window1, text="-engage early adopters for market validation", bg="gray4", fg="white",font="font 12").grid(row=11)
        Label(window1, text="-continually test with smaller, faster iterations", bg="gray4", fg="white",font="font 12").grid(row=12)
        Label(window1, text="-build a function, measure customer response, and verify/refute the idea", bg="gray4", fg="white",font="font 12").grid(row=13)
        Label(window1, text="-evidence-based decisions on when to pivot by changing your plan's course", bg="gray4", fg="white",font="font 12").grid(row=14)
        Label(window1, text="-maximize the efforts for speed, learning, and focus", bg="gray4", fg="white",font="font 12").grid(row=15)
        Label(window1, text="Market validation", bg="gray4", fg="white",font="font 14 bold").grid(row=16)
        Label(window1, text="A key principle of startup is to validate the market need before providing a customer-centric product.", bg="gray4", fg="white", font="font 12").grid(row=17)
        Label(window1, text="Market validation can be done in a number of ways, including surveys, cold calling, email responses.", bg="gray4", fg="white", font="font 12").grid(row=18)
        Label(window1, text="Design thinking", bg="gray4", fg="white", font="font 14 bold").grid(row=19)
        Label(window1, text="Design thinking is used to understand the customers' need in an engaged manner.", bg="gray4", fg="white", font="font 12").grid(row=20)
        Label(window1, text=".Design thinking and customer development can be biased because they do not remove the risk of bias ",bg="gray4", fg="white", font="font 12").grid(row=21)
        Label(window1,text=", the type of information sought, and the interpretation of that information.",bg="gray4", fg="white", font="font 12").grid(row=22)
        Label(window1, text="Encouraging people to “consider the opposite” of whatever decision they are about to make tends to reduce biases.", bg="gray4", fg="white", font="font 12").grid(row=23)
        Label(window1, text="Decision-making under uncertainty", bg="gray4", fg="white", font="font 14 bold").grid(row=24)
        Label(window1, text="In startups, many decisions are made under uncertainty, and hence a key principle for startups is to be agile and flexible.", bg="gray4", fg="white", font="font 12").grid(row=25)
        Label(window1,text="Founders can embed options to design startups in flexible manners, so that the startups can change easily in future.",bg="gray4", fg="white", font="font 12").grid(row=26)
        Label(window1,text="Uncertainty can vary within-person  and between-person. ",bg="gray4", fg="white", font="font 12").grid(row=27)
        Label(window1,text="A study found that when entrepreneurs feel more uncertain, they identify more opportunities.",bg="gray4", fg="white", font="font 12").grid(row=28)
        Label(window1,text="Partnering",bg="gray4", fg="white", font="font 14 bold").grid(row=29)
        Label(window1, text="Startups may form partnerships with other firms to enable their business model to operate. ", bg="gray4", fg="white", font="font 12").grid(row=30)
        Label(window1,text="To become attractive to other businesses, startups need to align their internal features.",bg="gray4", fg="white", font="font 12").grid(row=31)
        Label(window1,text="In their 2013 study, Kask and Linton develop two ideal profiles, or also known as configurations or archetypes",bg="gray4", fg="white", font="font 12").grid(row=32)
        Label(window1,text="The inheritor profile calls for a management style that is not too entrepreneurial and the startup should have an incremental invention.",bg="gray4", fg="white", font="font 12").grid(row=33)
        Label(window1,text="This profile is set out to be more successful in a market that has a dominant design.",bg="gray4", fg="white", font="font 12").grid(row=34)
        Label(window1,text="Startups usually need many different partners to realize their business idea.",bg="gray4", fg="white", font="font 12").grid(row=35)
        Label(window1, text="The commercialization process is often a bumpy road with iterations and new insights during the process. ", bg="gray4",fg="white", font="font 12").grid(row=36)
        Label(window1,text="Hasche and Linton (2018) argue that startups can learn from their relationships with other firms, and even if the relationship ends.",bg="gray4", fg="white", font="font 12").grid(row=37)
        Label(window1,text="When a relationship is failing for a startup it needs to make changes.",bg="gray4", fg="white", font="font 12").grid(row=38)


window = Tk()
window.configure(width=500, height=500, background="gray4")
Label(window, text="Welcome to the information menu ", bg="gray4", fg="white", font=" none 12 bold").grid\
    (row=0, column=0)
Label(window, text="Choose an option to see the information about:",bg="gray4",fg="white",font=" none 12 "
                                                                                               "bold").grid\
    (row=1, column=0)
options = ["Covid19 ", " guidelines", "vaccination", "mental health", "blood donation", "professional help",
           "system", "developer company"]
clicked = StringVar()
clicked.set(options[0])
drop = OptionMenu(window, clicked, *options)
drop.grid(row=2, column=0)
Button(window, text="view info", width=9, command=show_info).grid(row=3, column=0)
window.mainloop()
