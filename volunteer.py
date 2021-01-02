import openpyxl
from PIL import Image

print("Dear volunteer , we want you to be safe!")
print("here is a map, every city with a black star has a volunteer place:")
img = Image.open('map.jpg')
img.show()
print("Please sign in to one of the volunteering places :")
wb = openpyxl.load_workbook('volunteer.xlsx')
print("Enter the city you want to volunteer in:")
city = input()
sheet = wb[city]
print("Enter your name:")
name = input()
print("Enter your phone number:")
phone = input()
print("Enter the date you want to volunteer:")
date = input()
row1 = 0
for cell in sheet["A"]:    # find the empty raw
    if cell.value is None:
        row1 = cell.row
        break
    else:
        row1 = cell.row + 1

if row1 < 20:
    sheet.cell(row=row1, column=1).value = name
    sheet.cell(row=row1, column=2).value = phone
    sheet.cell(row=row1, column=3).value = date
    wb.save('volunteer_update.xlsx')
if row1 >= 20:
    print("there is no place left, try other time")
print("Thank you!")
