import openpyxl


def find_row():
    i = sheet.max_row
    return i


def info_input():
    last_row = find_row()
    print("enter the number of the test :")
    num_of_test = input()
    worksheet.cell(row=last_row, column=1).value = num_of_test

    print("enter your id:")
    work_id = input()
    worksheet.cell(row=last_row, column=2).value = work_id

    print("enter patient's id:")
    patient_id = input()
    worksheet.cell(row=last_row, column=3).value = patient_id

    print("enter the test time:")
    time = input()
    worksheet.cell(row=last_row, column=4).value = time

    print("enter test date:")
    date = input()
    worksheet.cell(row=last_row, column=5).value = date

    print("enter the test result:")
    test_res = input()
    worksheet.cell(row=last_row, column=6).value = test_res

    workbook.save()

    print("thank you!")


wb = openpyxl.load_workbook('tests_info.xlsx')  # for lab workers
print("dear worker ,to trace and find mistakes we need you enter the following right after finishing a test ")
print("enter the city you tested in: ")
city = input()
sheet = wb[city]    # the sheets are by cities
info_input()
