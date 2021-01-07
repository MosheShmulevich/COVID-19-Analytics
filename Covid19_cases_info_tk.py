from tkinter import *
import webbrowser
import openpyxl


def open_map():
    webbrowser.open('https://www.govmap.gov.il/sites/coronamap.html')


def by_city():
    city = clicked.get()

    file_name = 'confirmed_patients.xlsx'
    wb = openpyxl.load_workbook(file_name)
    sheet = wb[city]
    last_row = sheet.max_row
    confirmed_count = last_row - 1

    file_name1 = 'recovered.xlsx'
    wb = openpyxl.load_workbook(file_name1)
    sheet = wb[city]
    last_row1 = sheet.max_row
    recovered_count = last_row1 - 1

    file_name2 = 'death_cases.xlsx'
    wb = openpyxl.load_workbook(file_name2)
    sheet = wb[city]
    last_row2 = sheet.max_row
    death_count = last_row2 - 1

    Label(window, text="Total cases: ", bg="gray4", fg="white", font=" none 12 bold").grid(row=6, column=0)
    Label(window, text=confirmed_count, bg="gray4", fg="white", font="none 12 ").grid(row=7, column=0)
    Label(window, text="Recovered: ", bg="gray4", fg="white", font=" none 12 bold").grid(row=8, column=0)
    Label(window, text=recovered_count, bg="gray4", fg="white", font="none 12 ").grid(row=9, column=0)
    Label(window, text="Death Cases: ", bg="gray4", fg="white", font=" none 12 bold").grid(row=10, column=0)
    Label(window, text=death_count, bg="gray4", fg="white", font="none 12 ").grid(row=11, column=0)


window = Tk()
window.title("Covid19 Cases Info")
window.configure(width=200,height=200,background="gray4")
Label(window, text="Welcome to Covid19 Cases Info! ",bg="gray4",fg="white",font=" none 12 bold").grid(row=0, column=0)
Label(window, text="you can view the information by map ",bg="gray4",fg="white",font=" none 12 bold").grid(row=1, column=0,sticky=W)
Label(window, text="or by choosing a city",bg="gray4",fg="white",font=" none 12 bold").grid(row=2, column=0)
Button(window, text="Map",width=6,command = open_map).grid(row=3, column=0,sticky=W)
cities = ["Beer Sheva", "Tel Aviv", "Mizpe Ramon", "Heifa", "Afula", "Hadera", "Jerusalem", "Eilat"]
clicked = StringVar()
clicked.set(cities[0])
drop = OptionMenu(window, clicked, *cities)
drop.grid(row=3, column=1)
Button(window,text="Check",width=6,command= by_city).grid(row=5,column=1)


window.mainloop()

