from tkinter import *
from openpyxl import *
from tkinter import ttk

Patients_Details=load_workbook(filename='Patients_Details.xlsx',data_only=True)

class Statistic_Window(Tk):
    def __init__(self):
        super(Statistic_Window,self).__init__()
        self.title("Statistic Analysis")
        self.geometry("1000x500")
        self.minsize(1000,500)
        self.maxsize(1000,500)
        self.ComboBox_Statistic()
        self.configure(background="spring green")

################### create text box where the patient's details will be written.

    def text_box(self):
        text_=Text(self,width=100,height=400,)
        Text.pack(pady=20)
        button_frame=Frame(self)
        button_frame.pack()

        clear_button=Button(button_frame,text='Clear')
        clear_button.grid(row=0,column=0)
    ##############
    def create_message_tab(self):
        self.tab_control = ttk.Notebook(self)

        self.create_message_tab = ttk.Frame(self.tab_control)
        self.tab_control.add(self.create_message_tab, text="message_tab")

        self.Data = Text(self.LocationTab, state=DISABLED, font=(12))
        self.Data.place(x=500, y=20, width=600, height=500)

        self.DataScroll = ttk.Scrollbar(self.LocationTab, orient="vertical", command=self.Data.yview)
        self.DataScroll.place(x=1100, y=20, height=500)

    def close_text(self):
        if len(self.data.get("1.0","end-1c"))!=0:
            self.ClearLocation()

################### create options window to choose from.

    def switch(event, clicked):
        First_name = []
        Last_name= []
        ID= []
        Birth_Date= []
        Test_Date= []
        Patient_Status= []


        N=Patients_Details.nrows

        ################### printing the patient's details in the textbox window.

        if clicked.get()=='Firstname':

            for N in range (1,Patients_Details['גיליון1'].cell(row=N,column=0).value):
                print(First_name)

        elif clicked.get()=='Last name':
            for N in range (1,Patients_Details['גיליון1'].cell(row=N,column=1).value):
                print(Last_name)

        elif clicked.get() == 'ID':
            for N in range (1,Patients_Details['גיליון1'].cell(row=N,column=2).value):
                print(ID)

        elif clicked.get() == 'Birth Date':
            for N in range(1,Patients_Details['גיליון1'].cell(row=N,column=3).value):
                print(Birth_Date)

        elif clicked.get() == 'Test Date':
            for N in range (1,Patients_Details['גיליון1'].cell(row=N,column=4).value):
                print(Test_Date)

        elif clicked.get() == 'Patient Status':
            for N in range (1,Patients_Details['גיליון1'].cell(row=N,column=5).value):
                print(Patient_Status)

    ################### visual integrity,what the consumer will meet when the program runs.

    def ComboBox_Statistic(self):
        options = [
               "First name",
                "Last name",
                "ID",
                "Birth Date",
                "Test Date",
                "Patient Status",
        ]
        clicked=StringVar()
        clicked.set(options[0])

        drop=OptionMenu(Tk.self,clicked,*options,command=self.selected)
        drop.pack(pady=20)

        myButton= Button(Tk.self,text="Enter a statistical analysis:\n",command=self.selected)
        myButton.pack()

K=Statistic_Window()
K.mainloop()